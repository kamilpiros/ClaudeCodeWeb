"""Gemeinsame Bausteine zum Lesen einer DF_STATS Arbeitsmappe.

Wird von update_from_workbook.py benutzt. Hier steht alles, was direkt aus der
Mappe kommt: das Oeffnen trotz eingebetteter Zeichnungen, der Bilanzblock, das
Bussenjournal und die Anwesenheitsmatrix.
"""

import collections
import datetime
import io
import re
import zipfile

import openpyxl

# Die Mappe kennt MSO und MST als dieselbe Person, die Seite fuehrt sie als
# ein Mitglied. JWI war nur im Vereinsjahr 2017 dabei.
ALIAS = {"MSO": "MST/MSO", "MST": "MST/MSO", "MSO/MST": "MST/MSO"}

# Beschriftungen im Bussenjournal auf die Kategorien der Seite abbilden.
# Die Reihenfolge zaehlt: "Uninformed Absence (LIFO+)" und
# "LIFO+ / Uninformed Absence" gehoeren zu LIFO+, nicht zu Unentschuldigt,
# darum steht LIFO+ vor uninformed. SCHNIFO ist eine Eigenschoepfung des
# Statistikers und wird als LIFO gefuehrt.
REASON = [
    (r"^yearly", "Mitgliederbeitrag"),
    (r"^missing", "Absenz"),
    (r"lifo\+", "LIFO+"),
    (r"lifo|schnifo", "LIFO"),
    (r"uninformed", "Unentschuldigt"),
    (r"^li\b|^li$", "Last In"),
]


def norm(code):
    """Kuerzel saeubern. Fehlerwerte wie #VALUE! sind keine Mitglieder."""
    if not isinstance(code, str):
        return None
    c = code.strip()
    if not re.fullmatch(r"[A-Za-zÄÖÜäöü][A-Za-zÄÖÜäöü/]{1,7}", c):
        return None
    return ALIAS.get(c.upper(), c)


def reason_group(text):
    t = (text or "").strip().lower()
    for pat, name in REASON:
        if re.search(pat, t):
            return name
    return "Übrige"


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        m = re.match(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{2,4})\s*$", v)
        if m:
            d, mo, y = map(int, m.groups())
            return datetime.date(y + 2000 if y < 100 else y, mo, d)
    return None


def open_workbook(path):
    """Laedt die Mappe, nachdem Zeichnungen und Medien entfernt wurden.

    openpyxl stolpert sonst ueber die eingebetteten Grafiken
    (ValueError: Max value is 52).
    """
    drop = ("xl/drawings/", "xl/charts/", "xl/media/", "xl/embeddings/",
            "xl/ctrlProps/", "xl/activeX/")
    buf = io.BytesIO()
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name.startswith(drop):
                continue
            data = zin.read(name)
            if name.endswith(".rels"):
                data = re.sub(
                    rb'<Relationship[^>]*Target="[^"]*'
                    rb'(drawings|charts|media|embeddings|ctrlProps|activeX)/[^"]*"[^>]*/>',
                    b"", data)
            elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
                data = re.sub(rb"<drawing [^>]*/>", b"", data)
                data = re.sub(rb"<legacyDrawing [^>]*/>", b"", data)
                data = re.sub(rb"<controls>.*?</controls>", b"", data, flags=re.S)
                data = re.sub(rb"<oleObjects>.*?</oleObjects>", b"", data, flags=re.S)
            elif name == "[Content_Types].xml":
                data = re.sub(
                    rb'<Override[^>]*PartName="/xl/'
                    rb'(drawings|charts|media|embeddings|ctrlProps|activeX)/[^"]*"[^>]*/>',
                    b"", data)
            zout.writestr(name, data)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)


def year_sheets(wb):
    out = []
    for name in wb.sheetnames:
        m = re.match(r"^Stats (\d{4})$", name.strip())
        if m:
            out.append((int(m.group(1)), name))
    return sorted(out)


# --------------------------------------------------------------------------- #
# Kopfzeilen finden
# --------------------------------------------------------------------------- #
def _header(ws, first, second):
    """Kopfzeile suchen, in der links `first` und rechts davon `second` steht.

    Der Anker ist `second`, weil "Date:" im Blatt zweimal vorkommt: einmal
    ueber der Anwesenheit und einmal ueber dem Bussenjournal. Gesucht wird
    darum die naechstgelegene Datumsspalte links von `second`.
    """
    for r in range(1, 40):
        b = None
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower().startswith(second):
                b = c
                break
        if not b:
            continue
        a = None
        for c in range(b - 1, 0, -1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower().startswith(first):
                a = c
                break
        if not a:
            continue
        cols = []
        for c in range(b, b + 12):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str) or not v.strip():
                break
            cols.append(c)
        if cols:
            return r, a, cols
    return None


def members_row(ws, cols, hint=None):
    """Mitgliederkuerzel stehen ein paar Zeilen ueber der Kopfzeile.

    Zwei Stolpersteine: im Blatt 2018 hat die siebte Bussenspalte (JWI) keinen
    Eintrag in der Kuerzelzeile, und im Blatt 2026 steht dort ein #VALUE!.
    Darum gewinnt die Zeile mit den meisten Treffern, und wenn die Spaltenzahl
    zur Anwesenheit passt, werden Luecken aus deren Reihenfolge gefuellt.
    """
    best = {}
    for r in range(14, 24):
        got = {}
        for c in cols:
            n = norm(ws.cell(row=r, column=c).value)
            if n and len(n) <= 8 and not n.endswith(":"):
                got[c] = n
        if len(got) == len(cols):
            return got
        if len(got) > len(best):
            best = got
    if hint and len(hint) == len(cols):
        for i, c in enumerate(cols):
            best.setdefault(c, hint[i])
    return best if len(best) >= 4 else {}


# --------------------------------------------------------------------------- #
# Anwesenheit
# --------------------------------------------------------------------------- #
def attendance(ws):
    """[(datum, {mitglied: True/False}), ...] fuer jeden Termin des Blattes."""
    h = _header(ws, "date", "part")
    if not h:
        return [], {}
    hr, dc, cols = h
    who = members_row(ws, cols)
    if not who:
        return [], {}
    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        d = parse_date(ws.cell(row=r, column=dc).value)
        if not d:
            continue
        rec, gefuellt = {}, False
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                # Aeltere Blaetter haengen dem Wert eine Sortierziffer an,
                # 1.07 heisst anwesend, 0 heisst abwesend.
                rec[who[c]] = v >= 0.5
                gefuellt = True
            else:
                # Leere Zelle heisst abwesend. Ohne das fehlt der Termin in
                # Kopf an Kopf und in der Saisonauswertung.
                rec[who[c]] = False
        if gefuellt:
            rows.append((d, rec))
    return rows, who


# --------------------------------------------------------------------------- #
# Bussen
# --------------------------------------------------------------------------- #
def fines(ws, hint=None):
    """(je Mitglied und Grund, je Datum und Mitglied, Summe je Mitglied laut Blatt).

    `hint` ist die Mitgliederreihenfolge aus der Anwesenheitsmatrix und dient
    dazu, unlesbare Kuerzelzellen zu fuellen.
    """
    h = _header(ws, "date", "reason")
    if not h:
        return {}, [], {}
    hr, dc, cols = h
    rc = cols[0]
    cols = cols[1:]
    who = members_row(ws, cols, hint)

    by_reason = collections.defaultdict(lambda: collections.defaultdict(float))
    by_date = []
    for r in range(hr + 1, ws.max_row + 1):
        d = parse_date(ws.cell(row=r, column=dc).value)
        grund = reason_group(ws.cell(row=r, column=rc).value)
        rec = {}
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and v:
                m = who.get(c)
                if not m:
                    continue
                by_reason[m][grund] += v
                rec[m] = rec.get(m, 0) + v
        if d and rec:
            by_date.append((d, rec))

    # Zeile mit den Jahressummen je Mitglied, ein paar Zeilen ueber der Kopfzeile.
    # Sie enthaelt auch die nur pauschal gebuchten Bussen.
    total = {}
    for r in range(hr - 8, hr):
        got = {}
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and who.get(c):
                got[who[c]] = round(float(v), 2)
        if len(got) > len(total):
            total = got
    return ({m: dict(v) for m, v in by_reason.items()}, by_date, total)
