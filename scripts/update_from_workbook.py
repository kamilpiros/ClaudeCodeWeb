#!/usr/bin/env python3
"""Führt doenerstapler/df-assets.js aus einer DF_STATS Arbeitsmappe nach.

    python scripts/update_from_workbook.py DF_STATS_2026_08_07.xlsx

Was das Skript aus der Mappe zieht:

  years         Debitoren, Reserve, Rückstellung und Total je Vereinsjahr
  fiscalYears   erste und letzte Buchung je Jahresblatt, Anzahl Termine
  ledger        aufgelaufene Debitoren an jedem einzelnen Dönerfriitig
  timeline      der Stand dieser Mappe als weiterer Punkt
  spending      Ausgaben je Vereinsjahr, abgeleitet
  asOf/source   Stichtag und Dateiname

Was es bewusst nicht anfasst, weil es von Hand belegt wurde und in keiner Mappe
maschinenlesbar steht: crypto.marks, gold.marks, events, reserveSplit2020,
bench und gold.series. Diese Blöcke werden unverändert übernommen.

Der Goldkurs wird auf Wunsch mit --gold aus den beiden offenen Datensätzen
nachgeführt, die auch die Seite nennt.

Bekannte Eigenheiten der Mappe, die hier abgefangen werden:

  * Das Blatt 2018 hat sieben Betragsspalten, die siebte heisst nicht
    "Amount:" sondern "J dä verräter".
  * Einzelne Daten stehen als Text "22.12.2017" statt als Datum.
  * Bei CBO und YMI wurden 2018 bis 2020 und 2022 die Bussen nur als
    Jahressumme gebucht. Dieser Rest wird gleichmässig über die Termine des
    Jahres verteilt, damit jede Jahreskurve auf dem Jahrestotal endet.
  * Das Blatt 2026 weist die Merchandise-Rückstellung separat aus, alle
    früheren rechnen sie in die Reserve hinein. Ausgewiesen wird immer brutto.
  * openpyxl stolpert über die eingebetteten Zeichnungen, darum wird die Datei
    vorher im Speicher von drawings, charts und media befreit.
"""

import argparse
import collections
import csv
import datetime
import io
import json
import os
import re
import sys
import urllib.request
import zipfile

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl fehlt:  pip install openpyxl")

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASSETS = os.path.join(HERE, "doenerstapler", "df-assets.js")
OZ = 31.1034768

GOLD_CSV = "https://raw.githubusercontent.com/datasets/gold-prices/main/data/monthly.csv"
FX_CSV = "https://raw.githubusercontent.com/datasets/exchange-rates/main/data/monthly.csv"


# --------------------------------------------------------------------------- #
# Mappe oeffnen
# --------------------------------------------------------------------------- #
def open_workbook(path):
    """Laedt die Mappe, nachdem Zeichnungen und Medien entfernt wurden."""
    drop = ("xl/drawings/", "xl/charts/", "xl/media/", "xl/embeddings/",
            "xl/ctrlProps/", "xl/activeX/")
    buf = io.BytesIO()
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name.startswith(drop):
                continue
            data = zin.read(name)
            if name.endswith(".rels"):
                t = data.decode("utf-8")
                t = re.sub(r'<Relationship[^>]*Target="[^"]*'
                           r'(drawings|charts|media|embeddings|ctrlProps|activeX)/[^"]*"[^>]*/>',
                           "", t)
                data = t.encode("utf-8")
            elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
                t = data.decode("utf-8")
                t = re.sub(r"<drawing [^>]*/>", "", t)
                t = re.sub(r"<legacyDrawing [^>]*/>", "", t)
                t = re.sub(r"<controls>.*?</controls>", "", t, flags=re.S)
                t = re.sub(r"<oleObjects>.*?</oleObjects>", "", t, flags=re.S)
                data = t.encode("utf-8")
            elif name == "[Content_Types].xml":
                t = data.decode("utf-8")
                t = re.sub(r'<Override[^>]*PartName="/xl/'
                           r'(drawings|charts|media|embeddings|ctrlProps|activeX)/[^"]*"[^>]*/>',
                           "", t)
                data = t.encode("utf-8")
            zout.writestr(name, data)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)


def year_sheets(wb):
    out = []
    for name in wb.sheetnames:
        m = re.match(r"^Stats (\d{4})$", name.strip())
        if m:
            out.append((int(m.group(1)), name))
    return sorted(out)


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        m = re.match(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{2,4})\s*$", v)
        if m:
            d, mo, y = map(int, m.groups())
            return datetime.date(y + 2000 if y < 100 else y, mo, d)
    return None


# --------------------------------------------------------------------------- #
# Bilanzblock
# --------------------------------------------------------------------------- #
def balance(ws):
    """Liest Debitoren, Reserve, Ruecktstellung und Total aus dem Kopf des Blattes."""
    rec = {}
    for r in range(1, 26):
        for c in range(1, 32):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            lab = v.strip().upper()

            def nxt(step=4):
                for k in range(1, step + 1):
                    x = ws.cell(row=r, column=c + k).value
                    if isinstance(x, (int, float)):
                        return x
                return None

            if lab.startswith("TOTAL ASSETS"):
                rec["total"] = nxt(6)
            elif lab.startswith("DEBITORS TOTAL"):
                rec["debitors"] = nxt()
            elif (("RESERVES" in lab and ("TOTAL" in lab or "PAID" in lab))
                  or lab.startswith("INVESTMENTS")):
                # Die Beschriftung wechselt von Jahr zu Jahr: CAESHRESERVES
                # TOTAL, CRYPTORESERVES TOTAL*, GOLDRESERVES TOTAL*,
                # Investments TOTAL* und im Blatt 2020 CAESHRESERVES + Paid.
                # Der Zusatz TOTAL oder PAID ist noetig, sonst schnappt die
                # Bedingung im Blatt 2021 die Zeile "Reserves 22" mit dem
                # Uebertrag ins Folgejahr.
                rec["reserves"] = nxt()
                rec["label"] = v.strip()
            elif "MERCHANDISE" in lab:
                rec["provision"] = nxt()
    lab = (rec.get("label") or "").upper()
    rec["kind"] = ("Gold" if "GOLD" in lab or lab.startswith("INVESTMENTS")
                   else "Krypto" if "CRYPTO" in lab
                   else "Bargeld" if "RESERVES" in lab
                   else None)
    # Blatt 2026 fuehrt die Ruecktstellung neben der Reserve, frueher steckte
    # sie darin. Fuer die Zeitreihe wird immer brutto ausgewiesen.
    res = rec.get("reserves") or 0
    prov = rec.get("provision") or 0
    total = rec.get("total")
    if total and rec.get("debitors") is not None and abs(rec["debitors"] + res - total) > 0.01:
        res += prov
    rec["reservesGross"] = round(res, 2) if res else None
    return rec


# --------------------------------------------------------------------------- #
# Bussenjournal
# --------------------------------------------------------------------------- #
def find_ledger(ws):
    """Sucht die Kopfzeile mit Date und Reason und die Betragsspalten daneben."""
    for r in range(1, 40):
        dc = rc = None
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                if v.strip().lower().startswith("date"):
                    dc = c
                elif v.strip().lower().startswith("reason"):
                    rc = c
        if dc and rc and rc > dc:
            amt = []
            for c in range(rc + 1, rc + 12):
                # Bis zur ersten leeren Kopfzelle, sonst faellt die siebte
                # Spalte im Blatt 2018 unter den Tisch.
                if ws.cell(row=r, column=c).value in (None, ""):
                    break
                amt.append(c)
            if amt:
                return r, dc, rc, amt
    return None


def ledger(ws, target):
    f = find_ledger(ws)
    if not f:
        return [], 0.0, 0
    hr, dc, _, amt = f
    per = collections.defaultdict(float)
    alle = set()
    for r in range(hr + 1, ws.max_row + 1):
        d = parse_date(ws.cell(row=r, column=dc).value)
        if d:
            alle.add(d)
        s = 0.0
        for c in amt:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                s += v
        if s and d:
            per[d.isoformat()] += s
    dates = sorted(per)
    if not dates:
        return [], 0.0, len(alle)
    dated = sum(per.values())
    rest = round((target if target is not None else dated) - dated, 2)
    run, ser = 0.0, []
    for i, d in enumerate(dates, 1):
        run += per[d]
        ser.append([d, round(run + rest * i / len(dates), 2)])
    return ser, rest, len(alle)


def fiscal_range(ws):
    """Erste und letzte Buchung des Blattes, ueber Bussen- und Teilnahmespalte.

    Der Jahresbeitrag wird am ersten Doenerfriitig des Vereinsjahrs gebucht und
    ist damit der verlaesslichste Anker fuer den Jahresbeginn. Im Blatt 2024
    steht sonst eine verirrte Nachbuchung vom 28.04.2023 ganz vorne.
    """
    ds, start = [], None
    f = find_ledger(ws)
    rc = f[2] if f else None
    for col in (13, 4):
        for r in range(18, ws.max_row + 1):
            d = parse_date(ws.cell(row=r, column=col).value)
            if not d:
                continue
            ds.append(d)
            if col == 13 and rc:
                grund = ws.cell(row=r, column=rc).value
                if isinstance(grund, str) and "yearly" in grund.lower():
                    if start is None or d < start:
                        start = d
    if not ds:
        return None, None, 0
    return (start or min(ds)).isoformat(), max(ds).isoformat(), len(set(ds))


# --------------------------------------------------------------------------- #
# Goldkurs
# --------------------------------------------------------------------------- #
def fetch_gold_series():
    def get(url):
        req = urllib.request.Request(url, headers={"user-agent": "doenerfriitig.ch"})
        with urllib.request.urlopen(req, timeout=60) as r:
            return r.read().decode("utf-8").splitlines()

    gold = {r[0]: float(r[1]) for r in list(csv.reader(get(GOLD_CSV)))[1:] if r and r[1]}
    fx = {}
    for r in csv.reader(get(FX_CSV)):
        if len(r) > 2 and r[1] == "Switzerland":
            fx[r[0][:7]] = float(r[2])
    out = []
    for m in sorted(gold):
        if m in fx:
            oz = gold[m] * fx[m]
            out.append({"m": m, "usd": round(gold[m], 2),
                        "chfOz": round(oz, 2), "chfG": round(oz / OZ, 2)})
    return out


# --------------------------------------------------------------------------- #
# Hauptlauf
# --------------------------------------------------------------------------- #
def load_assets():
    raw = open(ASSETS, encoding="utf-8").read()
    return json.loads(raw[raw.index("{"):raw.rindex("}") + 1])


def save_assets(a):
    open(ASSETS, "w", encoding="utf-8").write(
        "window.DF_ASSETS = " + json.dumps(a, ensure_ascii=False, separators=(",", ":")) + ";\n")


def main():
    ap = argparse.ArgumentParser(description="df-assets.js aus einer Arbeitsmappe nachfuehren")
    ap.add_argument("workbook", help="Pfad zur DF_STATS Datei")
    ap.add_argument("--gold", action="store_true", help="Goldkursreihe mitnachfuehren")
    ap.add_argument("--dry-run", action="store_true", help="nur zeigen, nichts schreiben")
    args = ap.parse_args()

    wb = open_workbook(args.workbook)
    a = load_assets()
    alt = {y["year"]: dict(y) for y in a["years"]}

    sheets = year_sheets(wb)
    if not sheets:
        sys.exit("Keine Blaetter im Muster 'Stats JJJJ' gefunden")

    years, fiscal, led, residual = [], [], {}, {}
    for fy, name in sheets:
        ws = wb[name]
        b = balance(ws)
        von, bis, termine = fiscal_range(ws)
        if b.get("debitors") is None:
            sys.exit(f"Blatt {name}: DEBITORS TOTAL nicht gefunden")
        deb = round(b["debitors"], 2)
        res = b.get("reservesGross")
        prov = round(b.get("provision") or 0, 2)
        total = round(deb + (res or 0), 2)
        years.append({"year": fy, "debitors": deb, "reserves": res,
                      "reserveKind": b["kind"] if res else None,
                      "provision": prov, "total": total,
                      "from": von, "to": bis, "meets": termine})
        fiscal.append({"fy": fy, "from": von, "to": bis, "meets": termine})
        ser, rest, _ = ledger(ws, deb)
        if ser:
            led[str(fy)] = ser
            residual[str(fy)] = rest

    a["years"] = years
    a["fiscalYears"] = fiscal
    a["ledger"] = led
    a["ledgerResidual"] = residual

    # Ausgaben je Jahr, nach der bekannten Mechanik:
    # einkassierte Debitoren minus dem, was neu in die Reserve floss.
    # Der Anschaffungswert steht in reserveCost, sonst erschiene die Rendite
    # faelschlich als Ausgabe.
    cost = a.get("reserveCost", {})
    spend = []
    for i, y in enumerate(years):
        nxt = years[i + 1]["year"] if i + 1 < len(years) else None
        c = cost.get(str(nxt)) if nxt else None
        if y["reserves"] is None or c is None:
            continue
        to_res = round(c - y["reserves"], 2)
        spend.append({"year": y["year"], "collected": y["debitors"],
                      "toReserve": to_res, "spent": round(y["debitors"] - to_res, 2)})
    if spend:
        a["spending"] = spend
        a["spentTotal"] = round(sum(x["spent"] for x in spend), 2)

    # Diese Mappe als weiteren Punkt in der Zeitreihe
    cur = years[-1]
    stand = cur["to"]
    tl = {p["d"]: p for p in a.get("timeline", [])}
    tl[stand] = {"d": stand, "fy": cur["year"], "deb": cur["debitors"],
                 "res": cur["reserves"] or 0, "kind": cur["reserveKind"] or "keine",
                 "total": cur["total"], "src": "abschluss"}
    a["timeline"] = sorted(tl.values(), key=lambda p: p["d"])

    a["asOf"] = stand
    a["source"] = os.path.basename(args.workbook)

    if args.gold:
        try:
            ser = fetch_gold_series()
            if len(ser) > 100:
                a["gold"]["series"] = ser
                a["gold"]["last"] = ser[-1]
        except Exception as err:                       # noqa: BLE001
            print(f"Goldkurs nicht erreichbar, bleibt beim alten Stand: {err}")

    print(f"Mappe    {os.path.basename(args.workbook)}")
    print(f"Stand    {stand}  (letzter nachgetragener Doenerfriitig)")
    print(f"Punkte   {sum(len(v) for v in led.values())} im Bussenjournal, "
          f"{len(a['timeline'])} in der Zeitreihe")
    print()
    print(f"{'FJ':>5} {'Debitoren':>11} {'Reserve':>10} {'Total':>10}   Veraenderung")
    for y in years:
        v = alt.get(y["year"])
        delta = ""
        if not v:
            delta = "neu"
        elif abs((v.get("total") or 0) - y["total"]) > 0.005:
            delta = f"Total {v.get('total')} -> {y['total']}"
        elif abs((v.get("debitors") or 0) - y["debitors"]) > 0.005:
            delta = f"Debitoren {v.get('debitors')} -> {y['debitors']}"
        print(f"{y['year']:>5} {y['debitors']:>11.2f} "
              f"{(y['reserves'] or 0):>10.2f} {y['total']:>10.2f}   {delta}")

    if args.dry_run:
        print("\nProbelauf, nichts geschrieben.")
        return

    save_assets(a)
    print(f"\ndf-assets.js geschrieben, {os.path.getsize(ASSETS)} Bytes.")
    print("Nicht vergessen: den Versionsparameter an df-assets.js in "
          "doenerstapler/vermoegen.html hochzaehlen.")


if __name__ == "__main__":
    main()
